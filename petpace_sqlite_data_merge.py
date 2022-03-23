from datetime import datetime, timedelta
import csv
import argparse
from statistics import mean
import sqlite3
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table
from time import sleep
import re

DT_FORMATS = [
    "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%d/%m/%Y %I:%M:%S %p",
    "%d-%b-%y %H:%M:%S", "%d-%b-%y %I:%M:%S %p"]


def round_time(dt, roundTo=60):
    "round time to nearest minute"
    seconds = (dt.replace(tzinfo=None) - dt.min).seconds
    rounding = (seconds+roundTo/2) // roundTo * roundTo
    return dt + timedelta(0,rounding-seconds,-dt.microsecond)


def parse_dt(dt_string):
    "parse a date string into a timestamp, then round to minute"
    timestamp = None
    rounded = None
    for format in DT_FORMATS:
        try:
            timestamp = datetime.strptime(dt_string, format)
            rounded = round_time(timestamp)
            break
        except Exception as e:
            pass
    if timestamp is None:
        raise ValueError("Unable to parse datetime string: ", dt_string)
    return timestamp, rounded


def merge_pos(rows):
    "merge several rows re position of doggo to find pos with longest duration"
    positions = Counter()
    for row in rows:
        duration = int(row['duration'])
        if row['position'].lower().startswith('lying'):
            positions['Lying'] += duration
        elif row['position'].lower().startswith('eating'):
            positions['Standing'] += duration
        else:
            positions[row['position']] += duration
    c = positions.most_common(1)
    if c == []:
        return None, None
    else:
        return c[0]


def merge_activity(rows):
    "merge several rows re activity of doggo"
    activity = sum([row['activity'] for row in rows])
    activity_group = '\n'.join(set([row['activity_group'] for row in rows]))
    if activity_group == '':
        activity = None
    return activity, activity_group


def fetch(rows, fields):
    if len(rows) == 1:
        return [rows[0][field] for field in fields]
    elif len(rows) > 1:
        raise ValueError('several values for this timestamp:')
    else:
        return [None for field in fields]


def iter_timestamp(start, end, interval=timedelta(minutes=1), skip_times=[]):
    """iterate in one minute intervals between two datetimes,
    excluding specified periods"""
    timestamp = start
    while timestamp < end:
        if not any([skipstart <= timestamp < skipend for skipstart, skipend in skip_times]):
            yield timestamp
        timestamp += interval

def build_prox_table(headings):
    new_headings = [h.split('(')[1].split(')')[0] + " integer," for h in headings[1:]]
    query_fragment = ' '.join(new_headings).strip(',')
    query = f"""CREATE TABLE proximity (time timestamp,
            rounded_time timestamp, {query_fragment})"""
    return query

def build_sqlite(data_file, prox_file, datatable_file):
    """compile sqlite db in memory based on csv files"""
    con = sqlite3.connect(":memory:", detect_types=sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES)
    cur = con.cursor()
    cur.execute("""CREATE TABLE position (time timestamp,
    rounded_time timestamp, position text, duration text)""")
    cur.execute("""CREATE TABLE pulse (time timestamp,
    rounded_time timestamp, pulse integer)""")
    cur.execute("""CREATE TABLE respiration (time timestamp,
    rounded_time timestamp, respiration integer)""")
    cur.execute("""CREATE TABLE vvti (time timestamp,
    rounded_time timestamp, vvti real)""")
    cur.execute("""CREATE TABLE activity (time timestamp,
    rounded_time timestamp, activity real, activity_group text)""")
    cur.execute("""CREATE TABLE vector_mag (time timestamp,
    rounded_time timestamp, vector_mag real)""")
    con.commit()
    con.row_factory = sqlite3.Row

    with open(data_file, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for x in range(2):
            next(reader)
        fieldnames = next(reader)
        for row in reader:
            dict_row = dict(zip(fieldnames, row))
            timestamp, rounded = parse_dt(dict_row['Time'])
            if dict_row['Data type'] == 'Position':
                cur.execute(
                    "INSERT INTO position VALUES (?, ?, ?, ?)",
                    (timestamp, rounded, dict_row['Position'], dict_row['Position Duration']))
            elif dict_row['Data type'] == 'VVTI':
                cur.execute(
                    "INSERT INTO vvti VALUES (?, ?, ?)",
                    (timestamp, rounded, dict_row['VVTI']))
            elif dict_row['Data type'] == 'Respiration':
                cur.execute(
                    "INSERT INTO respiration VALUES (?, ?, ?)",
                    (timestamp, rounded, dict_row['Respiration']))
            elif dict_row['Data type'] == 'Pulse':
                cur.execute(
                    "INSERT INTO pulse VALUES (?, ?, ?)",
                    (timestamp, rounded, dict_row['Pulse']))
            elif dict_row['Data type'] == 'Activity':
                cur.execute(
                    "INSERT INTO activity VALUES (?, ?, ?, ?)",
                    (timestamp, rounded, dict_row['Activity'], dict_row['Activity Group']))
    con.commit()
    with open(prox_file, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        cur.execute(build_prox_table(reader.fieldnames))
        num_cols = len(reader.fieldnames[1:])
        for row in reader:
            timestamp, rounded = parse_dt(row['Timestamp'])
            cur.execute(
                f"INSERT INTO proximity VALUES (?, ?, {', '.join(['?' for x in range(num_cols)])})",
                (timestamp, rounded, *list(row.values())[1:]))
    with open(datatable_file, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for x in range(10):
            next(reader)
        fieldnames = next(reader)
        fieldnames = [f.strip() for f in fieldnames]
        for row in reader:
            dict_row = dict(zip(fieldnames, row))
            ts = f"{dict_row['Date']} {dict_row['Time']}"
            timestamp, rounded = parse_dt(ts)
            cur.execute(
                "INSERT INTO vector_mag VALUES (?, ?, ?)",
                (timestamp, rounded, dict_row['Vector Magnitude']))
    con.commit()
    return con


def merge_prox_rows(heads, rows):
    "merge several rows re proximity of doggo"
    data = {}
    for row in rows:
        cols = row[2:]
        for col_num, col in enumerate(cols):
            if data.get(col_num) is None:
                data[col_num] = dict(dist=[], present=[])
            try:
                original = int(col)
                dist = (((0.0012*(original**2))+(0.0936*original)+(1.9262)))
                data[col_num]['dist'].append(dist)
                data[col_num]['present'].append(1)
            except ValueError:
                data[col_num]['present'].append(0)
    for col, vals in data.items():
        if vals['dist'] != []:
            vals['dist'] = mean(vals['dist'])
        else:
            vals['dist'] = 0
        if vals['present'] != []:
            vals['present'] = mean(vals['present'])
        else:
            vals['present'] = 0
    flat_list = []
    for item in data.values():
        for val in item.values():
            flat_list.append(val)
    return flat_list


def main(data_file, prox_file, datatable_file, output, start, end, skip_rows=0, skip_times=[]):
    "do all the things"
    start_time = datetime.now()
    wb = Workbook()
    date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY HH:MM')
    RAW_sheet = wb.create_sheet('RAW')
    act_sheet = wb.create_sheet('Activity')
    act_sheet.append(['Time', 'Activity', 'Activity Group'])
    pos_dur_sheet = wb.create_sheet('PositionDuration')
    pos_dur_sheet.append(['Time', 'Position', 'Duration'])
    pulse_sheet = wb.create_sheet('Pulse')
    pulse_sheet.append(['Time', 'Pulse'])
    respiration_sheet = wb.create_sheet('Respiration')
    respiration_sheet.append(['Time', 'Respiration'])
    vvti_sheet = wb.create_sheet('VVTI')
    vvti_sheet.append(['Time', 'VVTI'])
    con = build_sqlite(data_file, prox_file, datatable_file)
    cursor = con.cursor()
    prox_sheet = wb.create_sheet('Proximity')
    cursor.execute("SELECT * FROM proximity")
    prox_heads = [x[0] for x in cursor.description]
    prox_sheet_heads = []
    for h in prox_heads[2:]:
        prox_sheet_heads.append(f'{h} Distance/min')
        prox_sheet_heads.append(f'{h} Present/min')
    prox_sheet.append(
        [
            'Time/min', *prox_sheet_heads, 'Activity', 'Pulse',
            'Respiration', 'VVTI', 'Position', 'Vector Magnitude'])

    with open(data_file, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            RAW_sheet.append(row)
    con = build_sqlite(data_file, prox_file, datatable_file)
    cursor = con.cursor()
    for timestamp in iter_timestamp(start, end, skip_times=skip_times):
        cursor.execute("SELECT rounded_time, position, duration FROM position WHERE rounded_time = '%s'" % timestamp)
        position, duration = merge_pos(cursor.fetchall())
        pos_dur_sheet.append([timestamp, position, duration])
        cursor.execute("SELECT rounded_time, activity, activity_group FROM activity WHERE rounded_time = '%s'" % timestamp)
        activity, activity_group = merge_activity(cursor.fetchall())
        act_sheet.append([timestamp, activity, activity_group])
        cursor.execute("SELECT rounded_time, pulse FROM pulse WHERE rounded_time = '%s'" % timestamp)
        pulse = fetch(cursor.fetchall(), ['pulse'])[0]
        pulse_sheet.append([timestamp, pulse])
        cursor.execute("SELECT rounded_time, vvti FROM vvti WHERE rounded_time = '%s'" % timestamp)
        vvti = fetch(cursor.fetchall(), ['vvti'])[0]
        vvti_sheet.append([timestamp, vvti])
        cursor.execute("SELECT rounded_time, respiration FROM respiration WHERE rounded_time = '%s'" % timestamp)
        respiration = fetch(cursor.fetchall(), ['respiration'])[0]
        respiration_sheet.append([timestamp, respiration])
        cursor.execute("SELECT vector_mag FROM vector_mag WHERE rounded_time = '%s'" % timestamp)
        vec_mag = cursor.fetchone()
        if vec_mag is not None:
            vec_mag = vec_mag['vector_mag']
        cursor.execute("SELECT * FROM proximity WHERE rounded_time = '%s'" % timestamp)
        prox_sheet.append([timestamp, *merge_prox_rows(prox_heads, cursor.fetchall()), activity, pulse, respiration, vvti, position, vec_mag])
    for worksheet in wb.worksheets:
        for cell in worksheet['A'][1:]:
            cell.style = date_style
    wb.save(output)
    con.close()
    print(datetime.now() - start_time)

if __name__ == "__main__":
    if __name__ == '__main__':
        parser = argparse.ArgumentParser(
            description='script to merge a bunch of doggo data')
        parser.add_argument('data', metavar='i', help='input export file')
        parser.add_argument('proximity', metavar='p', help='input proximity sheet')
        parser.add_argument('datatable', metavar='p', help='input datatable sheet')
        parser.add_argument('output', metavar='o', help='output xlsx file')
        parser.add_argument(
            '--start', '-s', type=datetime.fromisoformat, help='start of monitoring period')
        parser.add_argument(
            '--end', '-e', type=datetime.fromisoformat, help='end of monitoring period')
        parser.add_argument(
            '--chargetimes', '-t', nargs='+', type=datetime.fromisoformat, default=[], help='time periods to skip')
        args = parser.parse_args()
        skiptimes = [(args.chargetimes[x-1], args.chargetimes[x]) for x in range(1, len(args.chargetimes), 2)]
        main(
            args.data, args.proximity, args.datatable, args.output, args.start, args.end, skip_times=skiptimes)
